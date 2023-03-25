from helper import crear_muebles,generar_rango_cruza,llenar_resultado,arr_numeros,crear_pedidos
import random
import matplotlib.pyplot as plt
import numpy as np
import tkinter as tk
from Interfaz import Interfaz
from openpyxl import Workbook

class AlgoritmoGenetico:
    def __init__(self, muebles,inventario,pedidos,n_individuos, tamanio_poblacion,n_generaciones,prob_mutacion,n_mutaciones,prob_mutacion_gen):
        self.tamanio_poblacion = tamanio_poblacion
        self.n_individuos = n_individuos
        self.muebles = muebles
        self.inventario = inventario
        self.pedidos = pedidos
        self.n_generaciones = n_generaciones
        self.poblacion=[]
        self.mejor_individuo=[]
        self.media_individuo=[]
        self.peor_individuo=[]
        self.prob_mutacion=prob_mutacion
        self.n_mutaciones=n_mutaciones
        self.prob_mutacion_gen=prob_mutacion_gen
        self.cant_genes=arr_numeros(self.n_individuos)
        self.penalizacion=100
        self.primera_genetica()
        self.bucle_algoritmo()
        self.generar_tabla()
    
    def generar_tabla(self):
        labels=['id','Mueble','Precio','Cantidad','Pata silla delantera','Pata silla trasera','Pata mesa','Refuerzos','Tornillos','Respaldos','Asientos','Tablas','Tuercas','Ganancia del pedido']
        wb = Workbook()
        sheet=wb.active
        for index,dato in enumerate(labels):
            sheet.cell(row=1, column=index+1, value=dato)
        mejor_individuo=self.mejor_individuo[-1]
        print(f'Individuo: \n{mejor_individuo}')
        for index,id in enumerate(mejor_individuo['data']):
            pedido=list(pedido for pedido in self.pedidos if pedido['id']  == id)[0]
            mueble=list(mueble for mueble in self.muebles if mueble.nombre == pedido['mueble'])[0]
            sheet.cell(row=index+2, column=1, value=id)
            sheet.cell(row=index+2, column=2, value=mueble.nombre)
            sheet.cell(row=index+2, column=3, value=mueble.precio)
            sheet.cell(row=index+2, column=4, value=pedido['cantidad'])
            sheet.cell(row=index+2, column=14, value=pedido['ganancia'])
        sheet.cell(row=len(mejor_individuo['data'])+2, column=14, value='Total:')
        sheet.cell(row=len(mejor_individuo['data'])+3, column=14, value=mejor_individuo['ganancia'])
        for index,arr in enumerate(mejor_individuo['componentes']):
            for j,value in enumerate(arr):
                sheet.cell(row=index+2, column=j+5, value=value)
        for index,value in enumerate(self.inventario.values()):
            sheet.cell(row=len(mejor_individuo['data'])+2, column=index+5, value=value)
        for index in range(14):
            sheet.cell(row=len(mejor_individuo['data'])+2, column=index+1, value='----------------')
            
        sheet.cell(row=len(mejor_individuo['data'])+3, column=3, value='Inventario')
        sheet.cell(row=len(mejor_individuo['data'])+3, column=4, value='Inicial')
        for index,value in enumerate(mejor_individuo['inventario']):
            sheet.cell(row=len(mejor_individuo['data'])+3, column=index+5, value=value)
        sheet.cell(row=len(mejor_individuo['data'])+4, column=3, value='Inventario')
        sheet.cell(row=len(mejor_individuo['data'])+4, column=4, value='Final')

        

        wb.save('Tabla_mejor_individuo.xlsx')
    
    def primera_genetica(self):
        individuo_aux=[]
        for pedido in self.pedidos:
            individuo_aux.append(pedido.get('id'))
        for _ in range(self.tamanio_poblacion):
            random.shuffle(individuo_aux)
            individuo = individuo_aux.copy()
            self.poblacion.append(self.crear_individuo(individuo))
        print(*self.poblacion,sep='\n')
    def bucle_algoritmo(self):
        aux = 0
        while (aux < self.n_generaciones):
            self.generacion=aux+1
            poblacion_nueva = []
            # Hacer todo el bucle del algoritmo
            for i in range(len(self.poblacion)):
                if i == (len(self.poblacion)-1):
                    individuo = self.cruza(self.poblacion[i].get(
                        'data'), self.poblacion[0].get('data'))
                    poblacion_nueva.append(individuo)
                else:
                    individuo = self.cruza(self.poblacion[i].get(
                        'data'), self.poblacion[i+1].get('data'))
                    poblacion_nueva.append(individuo)
            # Se hace la mutación
            self.mutacion(poblacion_nueva)

            # Se calcula la aptitud para cada individuo y se agrega id
            poblacion_aptitud=[]
            for indiv in poblacion_nueva:
                poblacion_aptitud.append(self.agregar_aptitud(indiv))
            # Se insertan los nuevos en la población general
            for indiv in poblacion_aptitud:
                self.poblacion.append(indiv)
            self.ordenar_poblacion_por_aptitud()
            # Poda hasta tener el numero de individuos iniciales
            self.graficar_individuos(aux+1)
            self.poda()
            self.mejor_individuo.append(self.poblacion[0])
            self.media_individuo.append(self.poblacion[round(self.tamanio_poblacion/2)])
            self.peor_individuo.append(self.poblacion[self.tamanio_poblacion-1])
            print(f'Mejor individuo: {self.poblacion[0]}')
            print(f'Peor individuo: {self.poblacion[self.tamanio_poblacion-1]}')
            print(f'Generación {aux+1}')
            for indiv in self.poblacion:
                print(indiv)
            aux += 1
    def mutacion(self, nueva_poblacion):
        for individuo in nueva_poblacion:
            if random.uniform(0,1) <= self.prob_mutacion:
                for _ in range(self.n_mutaciones):
                    if random.uniform(0,1) <= self.prob_mutacion_gen:
                        a, b = random.choices(self.cant_genes, k=2)
                        individuo.get('data')[a], individuo.get('data')[
                            b] = individuo.get('data')[b], individuo.get('data')[a]
    def poda(self):
        while len(self.poblacion) != self.tamanio_poblacion:
            self.poblacion.pop()
    def cruza(self, indiv1, indiv2):
        # llenado parametros iniciales
        resultado = []
        indiv1_copy = indiv1.copy()
        index_aux = 0
        # Se crea un arreglo temporal con -1 el cual será el arreglo que se retorna como resultado de la cruza
        resultado = llenar_resultado(self.n_individuos)
        a, b = generar_rango_cruza(self.cant_genes)

        # se introduce a resultado los valores seleccionados del primer individuo
        for i in range(a, b+1):
            resultado[i] = indiv1[i]
            index_aux = i

        # for para eliminar los datos que ya se usaron del individuo1 (se creó un arreglo auxiliar)
        for i in range(a, b+1):
            indiv1_copy.remove(indiv1[i])

        datos_restantes = len(indiv1)-len(indiv1[a:b+1])
        index_aux += 1
        aux_result = index_aux
        aux_indv2 = index_aux
        for _ in range(datos_restantes):

            band = False
            while band == False:
                if aux_indv2 >= len(indiv2):
                    aux_indv2 = 0
                if aux_result >= len(resultado):
                    aux_result = 0
                if indiv2[aux_indv2] in indiv1_copy:
                    resultado[aux_result] = indiv2[aux_indv2]
                    indiv1_copy.remove(indiv2[aux_indv2])
                    aux_result += 1
                    aux_indv2 += 1
                    band = True
                else:
                    aux_indv2 += 1
        individuo = self.crear_data(resultado)
        return individuo
    def agregar_aptitud(self, individuo):
        return self.calcular_data(individuo.get('data'))
    def ordenar_poblacion_por_aptitud(self):
        self.poblacion.sort(key=lambda aptitud: aptitud['ganancia'], reverse=True)
    def crear_data(self, data):
        return {'data':data}
    def calcular_data(self, data):
        pedidos_de_individuo=[]
        ganancia=0
        #Encuentro los los pedidos en orden del individuo
        for gen in data:
            pedido=list(pedido for pedido in self.pedidos if pedido['id']  == gen)[0]
            pedidos_de_individuo.append(pedido)

        inventario_copy=list(self.inventario.values()).copy()
        band=False
        #Itero cada pedido para realizar el decremento de material en inventario
        muebles=0
        arr_a_restar=[]
        for i,pedido in enumerate(pedidos_de_individuo):
            arr_pedido=list(pedido.values()) # Genero un arreglo con [id_pedido, cantidad_de_mueble, nombre_mueble]
            mueble=list(mueble for mueble in self.muebles if mueble.nombre == arr_pedido[2])[0] #Traigo el mueble del pedido
            arr_aux=mueble.get_material() #Obtengo los material que usa para crear el mueble
            arr_material=[]
            #Se multiplica cantidad de productos * c/componente de fabricacion
            for cantidad in arr_aux:
                arr_material.append(arr_pedido[1]*cantidad) #El resultado es un arreglo con todo el material usado para entregar el pedido
            arr_a_restar.append(arr_material)
            #Se evalua si se puede cumplir el pedido
            for index,cantidad in enumerate(arr_material):
                resta=inventario_copy[index]-cantidad
                if  resta<0:
                    band=True                    
                else:
                    inventario_copy[index]=resta
            if band==False:
                muebles+=1
                ganancia+=pedido['ganancia']
            else:
                inventario=self.calcular_inventario(arr_a_restar)
                break
        # return {'data':data,'ganancia':ganancia, 'muebles':muebles, 'inventario':inventario}
        return {'data':data,'ganancia':ganancia, 'muebles':muebles, 'inventario':inventario, 'componentes':arr_a_restar}

    def calcular_inventario(self, arr_a_restar):
        arr_a_restar.pop()
        inventario_copy2=list(self.inventario.values()).copy()
        for arr in arr_a_restar:
            for index,dato in enumerate(arr):
                inventario_copy2[index]=inventario_copy2[index]-dato
        return inventario_copy2

    def crear_individuo(self,individuo_data):
        data= self.calcular_data(individuo_data)
        individuo = data
        return individuo
    def graficar_individuos(self,n_generacion):
        fig,aux=plt.subplots()
        arr_x=[]
        arr_y=[]
        for index,individuo in enumerate(self.poblacion):
            arr_x.append(index+1)
            arr_y.append(individuo['ganancia'])
        plt.stem(arr_x,arr_y,label='Ganancia')
        aux.set_title(f'Individuos en generacion: {n_generacion}',fontdict={'fontsize':20,'fontweight':'bold'})
        aux.set_xlabel('Individuo',fontdict={'fontsize':15,'fontweight':'bold', 'color':'tab:red'})
        aux.set_ylabel('Ganancia',fontdict={'fontsize':15,'fontweight':'bold', 'color':'tab:blue'})
        aux.legend(loc='upper right',prop={'size':10})
        # plt.grid()
        plt.savefig(f'./images/generacion_{self.generacion}')
        plt.close()
def generar_grafica(algoritmo):

    list_epocas = []
    list_mejores_aptitud = []
    list_peores_aptitud = []
    list_media_aptitud=[]
    for x in algoritmo.media_individuo:
        list_media_aptitud.append(x.get('ganancia'))
    for k in algoritmo.mejor_individuo:
        list_mejores_aptitud.append(k.get('ganancia'))
    for j in algoritmo.peor_individuo:
        list_peores_aptitud.append(j.get('ganancia'))
    for i in range(algoritmo.n_generaciones):
        list_epocas.append(i+1)  
    fig, ax = plt.subplots()
    ax.plot(list_epocas, list_mejores_aptitud,label='Mejores Aptitud')
    ax.plot(list_epocas, list_media_aptitud,label='Aptitud Media')
    ax.plot(list_epocas, list_peores_aptitud, color='red',label='Peores Aptitud')
    # index=algoritmo.n_generaciones -1
    # ax.text(0.5,5050,f'Mejor individuo {algoritmo.mejor_individuo[index]}')
    ax.legend(loc='lower right')
    plt.savefig('images/evolucion_aptitud')
    plt.show()  



if __name__ == "__main__":
    # window = tk.Tk()
    # entrada= Interfaz(window)
#                          15            1                                  6                        1000                      300                          
    muebles=crear_muebles()
    # pedidos=crear_pedidos(7,muebles,5,15)
    pedidos=crear_pedidos(30,muebles,5,15)
    inventario={'pata_silla_delantera':200,'pata_silla_trasera':200,'pata_mesa':400,'refuerzos':400,'tornillos':600,'respaldos':50,'asientos':50,'tablas':100,'tuercas':600}

    AG=AlgoritmoGenetico(muebles,
                        n_individuos=len(pedidos), #15
                        pedidos=pedidos,
                        inventario=inventario,
                        # tamanio_poblacion= entrada.get_tamanio_pob(), #20 
                        # n_generaciones=entrada.get_generaciones(), #10
                        # prob_mutacion=entrada.get_prob_mutacion(), #0.7
                        # n_mutaciones=entrada.get_n_mutacion(), #2
                        # prob_mutacion_gen=entrada.get_prob_muta_gen() #0.7
                        tamanio_poblacion= 20 ,
                        n_generaciones=20,
                        prob_mutacion=0.7,
                        n_mutaciones= 2,
                        prob_mutacion_gen= 0.7
                        )
    generar_grafica(AG)